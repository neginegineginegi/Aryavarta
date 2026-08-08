#!/usr/bin/env python3
"""Transform three GEM industrial datasets into Abhilekh inbox rows.

Usage:
  python3 extract-gem-industry-2026.py --steel-plants A.xlsx --cement B.xlsx \
      --metcoal C.xlsx [--out DIR]

Datasets (all Global Energy Monitor, CC BY 4.0; raw files stay OUTSIDE the
repository, recorded in data/raw/gem/MANIFEST.md):

- Global Iron and Steel Tracker, June 2026 (V1), plant-level workbook.
  Emits nominal crude steel capacity of operating plants, by state and for
  India. Plant-level nominal capacity is a DIFFERENT measure from the
  unit-level furnace sums already published; each methodology names its own
  basis and neither is corrected toward the other.
- Global Cement and Concrete Tracker, July 2026 (V1). Emits operating cement
  capacity by state and for India.
- Production and Consumption of Met Coal and Iron Ore by Steel Industry,
  December 2025 (V1). Emits four national values (met coal mined, iron ore
  mined, pig iron produced, DRI produced). The file's "consumed" columns are
  pure factor-multiplication estimates and are deliberately not published.

Verdicts recorded for the other two files in this batch, with no output:
Portal Energetico (August 2026) holds zero India rows (Latin America portal);
the Global Chemicals Inventory (November 2025 V2) carries no numeric fields,
only product and feedstock text, so there is nothing to publish as a value.
"""

import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

OUT_IND = "id,name,unit,category,methodology\n"
OUT_VAL = (
    "indicator,state,year,value,source_title,source_url,reporting_period,"
    "reporting_org,notes,verified_on\n"
)
VERIFIED_ON = "2026-08-08"
GEM_URL = "https://globalenergymonitor.org/"


def q(s: str) -> str:
    return '"' + s.replace('"', '""') + '"'


def arg(name: str) -> str | None:
    return sys.argv[sys.argv.index(name) + 1] if name in sys.argv else None


def header_index(ws, required):
    rows = ws.iter_rows(values_only=True)
    for r in rows:
        if r and sum(1 for c in r if c is not None) > 3:
            hdr = [str(c) if c is not None else "" for c in r]
            for col in required:
                if col not in hdr:
                    sys.exit(f"FATAL: {ws.title}: expected column {col!r} missing")
            return {h: i for i, h in enumerate(hdr)}, rows
    sys.exit(f"FATAL: {ws.title}: no header row found")


def steel_plants(path: str, ind, val, report):
    cite = "Global Energy Monitor, Global Iron and Steel Tracker, June 2026 (V1) release"
    url = "https://globalenergymonitor.org/projects/global-iron-and-steel-tracker/"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    ix, rows = header_index(wb["Plant data"], ["GEM plant ID", "Country/area", "Subnational unit"])
    state_of = {}
    for r in rows:
        if r and str(r[ix["Country/area"]]).strip() == "India":
            state_of[str(r[ix["GEM plant ID"]])] = str(r[ix["Subnational unit"]]).strip()

    cap_col = "Nominal crude steel capacity (ttpa)"
    ix, rows = header_index(
        wb["Plant capacities and status"], ["GEM plant ID", "Country/area", "Status", cap_col]
    )
    by_state: dict[str, float] = defaultdict(float)
    n = skipped_state = 0
    for r in rows:
        if r is None or str(r[ix["Country/area"]]).strip() != "India":
            continue
        if str(r[ix["Status"]]).strip() not in ("operating", "operating pre-retirement"):
            continue
        cap = r[ix[cap_col]]
        if not isinstance(cap, (int, float)):
            continue
        st = state_of.get(str(r[ix["GEM plant ID"]]), "")
        if not st or st.lower() == "unknown":
            skipped_state += 1
            continue
        by_state[st] += float(cap)
        n += 1
    wb.close()

    slug = "steel-capacity-crude-nominal"
    method = (
        "Sum of nominal crude steel capacity over GEM-tracked Indian plants "
        "with status operating or operating pre-retirement, attributed to the "
        "state of the plant. Plant-level nominal capacity; the archive's "
        "unit-level steel indicators sum individual tracked furnaces instead, "
        "so the two measures differ and neither is adjusted toward the other. "
        "The tracker covers only plants of 0.5 million tonnes per annum and "
        f"greater. Snapshot at the named release. Source: {cite}, CC BY 4.0."
    )
    ind.append(f"{slug},{q('Steel capacity, nominal crude, operating plants')},ttpa,Industry,{q(method)}\n")

    def emit(state: str, v: float, note: str = ""):
        val.append(
            f"{slug},{q(state)},2026,{v:.0f},{q(cite)},{url},"
            f"{q('June 2026 (V1) release')},Global Energy Monitor,{q(note)},{VERIFIED_ON}\n"
        )

    total = sum(by_state.values())
    emit("India", total, "Sum over all tracked Indian plants, all states.")
    for st in sorted(by_state):
        emit(st, by_state[st])
    report.append(
        f"steel plants: {n} operating capacity rows over {len(by_state)} states, "
        f"total {total:,.0f} ttpa; rows without a state {skipped_state}"
    )


def cement(path: str, ind, val, report):
    cite = "Global Energy Monitor, Global Cement and Concrete Tracker, July 2026 (V1) release"
    url = "https://globalenergymonitor.org/projects/global-cement-and-concrete-tracker/"
    cap_col = "Cement capacity (million metric tonnes per annum)"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ix, rows = header_index(
        wb["Final data"],
        ["GEM plant ID", "Country/area", "Subnational unit", "Operating status", cap_col],
    )
    by_state: dict[str, float] = defaultdict(float)
    n = missing_cap = 0
    for r in rows:
        if r is None or str(r[ix["Country/area"]]).strip() != "India":
            continue
        if str(r[ix["Operating status"]]).strip() != "operating":
            continue
        cap = r[ix[cap_col]]
        if not isinstance(cap, (int, float)):
            missing_cap += 1
            continue
        by_state[str(r[ix["Subnational unit"]]).strip()] += float(cap)
        n += 1
    wb.close()

    slug = "cement-capacity-operating"
    method = (
        "Sum of cement capacity over GEM-tracked Indian cement plants with "
        "status operating, attributed to the state of the plant, per the "
        "tracker's own inclusion criteria. Snapshot at the named release, not "
        f"a historical series. Source: {cite}, CC BY 4.0."
    )
    ind.append(f"{slug},{q('Cement capacity, operating')},mtpa,Industry,{q(method)}\n")

    def emit(state: str, v: float, note: str = ""):
        val.append(
            f"{slug},{q(state)},2026,{v:.1f},{q(cite)},{url},"
            f"{q('July 2026 (V1) release')},Global Energy Monitor,{q(note)},{VERIFIED_ON}\n"
        )

    total = sum(by_state.values())
    emit("India", total, "Sum over all tracked Indian plants, all states.")
    for st in sorted(by_state):
        emit(st, by_state[st])
    report.append(
        f"cement: {n} operating plants with capacity over {len(by_state)} states, "
        f"total {total:,.1f} mtpa; operating plants without capacity {missing_cap}"
    )


def metcoal(path: str, ind, val, report):
    cite = (
        "Global Energy Monitor, Production and Consumption of Met Coal and Iron Ore "
        "by Steel Industry, December 2025 (V1) release"
    )
    url = "https://globalenergymonitor.org/projects/global-iron-and-steel-tracker/"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    cols = {
        "met-coal-mined": (
            "Met coal mined (ttpa)",
            "Metallurgical coal mined",
            "Country-level metallurgical coal production, estimated by GEM by "
            "summing mine-level production, using mine capacity where "
            "production is unknown.",
        ),
        "iron-ore-mined": (
            "Iron ore mined (ttpa)",
            "Iron ore mined",
            "Country-level iron ore production, estimated by GEM by summing "
            "mine-level production, applying a 45 percent utilization rate to "
            "capacity where production is unknown.",
        ),
        "pig-iron-produced": (
            "Pig iron produced (ttpa)",
            "Pig iron produced",
            "Country-level pig iron production as reported by the World Steel "
            "Association, with GEM blast furnace capacity used where WSA "
            "production is unknown.",
        ),
        "dri-produced": (
            "DRI produced (ttpa)",
            "Direct reduced iron produced",
            "Country-level DRI production as reported by the World Steel "
            "Association, with GEM DRI capacity used where WSA production is "
            "unknown.",
        ),
    }
    ix, rows = header_index(wb["ProdConsMetCoalIronOre"], ["Country"] + [c[0] for c in cols.values()])
    india = next((r for r in rows if r and str(r[ix["Country"]]).strip() == "India"), None)
    wb.close()
    if india is None:
        sys.exit("FATAL: metcoal: no India row")

    for slug, (col, name, basis) in cols.items():
        v = india[ix[col]]
        if not isinstance(v, (int, float)):
            report.append(f"metcoal: {col} not numeric for India, skipped")
            continue
        method = (
            f"{basis} The file's consumption columns are factor-based estimates "
            "and are not published here. Snapshot for the release period. "
            f"Source: {cite}, CC BY 4.0."
        )
        ind.append(f"{slug},{q(name)},ttpa,Industry,{q(method)}\n")
        val.append(
            f"{slug},India,2025,{v:.0f},{q(cite)},{url},"
            f"{q('December 2025 (V1) release')},Global Energy Monitor,,{VERIFIED_ON}\n"
        )
        report.append(f"metcoal India {slug}: {v:,.0f} ttpa")


def main() -> None:
    out = Path(arg("--out") or "out")
    out.mkdir(parents=True, exist_ok=True)
    ind: list[str] = []
    val: list[str] = []
    report: list[str] = []
    if arg("--steel-plants"):
        steel_plants(arg("--steel-plants"), ind, val, report)
    if arg("--cement"):
        cement(arg("--cement"), ind, val, report)
    if arg("--metcoal"):
        metcoal(arg("--metcoal"), ind, val, report)
    if not ind:
        sys.exit(__doc__)
    (out / "industry-indicators.csv").write_text(OUT_IND + "".join(ind))
    (out / "industry-indicator_values.csv").write_text(OUT_VAL + "".join(val))
    (out / "industry-report.txt").write_text("\n".join(report) + "\n")
    print("\n".join(report))
    print(f"\nwrote {out}/industry-indicators.csv, industry-indicator_values.csv")


if __name__ == "__main__":
    main()
