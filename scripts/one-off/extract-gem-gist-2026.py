#!/usr/bin/env python3
"""Transform GEM's Global Iron and Steel Tracker into Abhilekh inbox rows.

Usage: python3 extract-gem-gist-2026.py <path-to-GIST-xlsx> [--out DIR]

Raw file: Global Iron and Steel Tracker, June 2026 (V1) release, Global
Energy Monitor, CC BY 4.0. The raw workbook stays OUTSIDE the repository
(data/raw/ is gitignored); this script is the reproducible bridge from it to
the two committed inbox sheets.

Pipeline stages, per docs/DEVELOPMENT_DATA.md:
  validate (prove schema, count anomalies, refuse ambiguity)
  -> filter (India rows only)
  -> aggregate (unit rows to national snapshot indicators)
  -> emit (indicators.csv + indicator_values.csv fragments + report)

What is deliberately NOT produced: any year-by-year series. Start years are
missing for 77% of Indian units, including every basic oxygen furnace, so a
historical series would present a fifth of the industry as the whole. These
are snapshot indicators dated to the release.
"""

import hashlib
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

RELEASE = "June 2026 (V1)"
RELEASE_YEAR = 2026
CITATION = "Global Energy Monitor, Global Iron and Steel Tracker, June 2026 (V1) release"
SOURCE_URL = "https://globalenergymonitor.org/projects/global-iron-and-steel-tracker/"
VERIFIED_ON = "2026-08-08"  # date this extraction was run and reviewed

# Sheet name -> (route slug, route label). The sheet IS the classification.
SHEETS = {
    "Electric arc furnaces": ("eaf", "electric arc"),
    "Basic oxygen furnaces": ("bof", "basic oxygen"),
    "Induction furnaces": ("if", "induction"),
    "Open hearth furnaces": ("ohf", "open hearth"),
}

# The exact header cores proven during inspection; a mismatch means GEM
# changed the format and every assumption must be re-checked by a human.
CORE_HEADER = [
    "GEM plant ID", "GEM unit ID", "Unit name", "GEM wiki page", "Country/area",
    "Unit status", "Announced date", "Construction date", "Start date", "Unit age",
]
CAPACITY_COL = "Current capacity (ttpa)"

STATUSES = {
    "announced", "construction", "operating", "operating pre-retirement",
    "mothballed", "mothballed pre-retirement", "cancelled", "retired",
}
# Snapshot measures. "Operating" includes units operating while slated for
# retirement: they are producing steel today, and the distinction is stated
# in the methodology rather than silently dropped.
MEASURES = {
    "operating": {"operating", "operating pre-retirement"},
    "construction": {"construction"},
    "announced": {"announced"},
}


def fail(msg: str) -> None:
    print(f"FATAL: {msg}", file=sys.stderr)
    sys.exit(1)


def main() -> None:
    if len(sys.argv) < 2:
        fail(__doc__.splitlines()[2])
    raw = Path(sys.argv[1])
    out = Path(sys.argv[sys.argv.index("--out") + 1]) if "--out" in sys.argv else Path("out")
    out.mkdir(parents=True, exist_ok=True)

    sha = hashlib.sha256(raw.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(raw, data_only=True)

    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            fail(f"expected sheet {sheet!r} missing; sheets = {wb.sheetnames}")

    report: list[str] = [
        f"GIST extraction report",
        f"raw file: {raw.name} ({raw.stat().st_size} bytes)",
        f"sha256: {sha}",
        "",
    ]

    # validate + filter + collect ------------------------------------------
    india_units: list[dict] = []
    seen_ids: Counter = Counter()
    anomalies: Counter = Counter()
    for sheet, (route, _) in SHEETS.items():
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        if hdr[: len(CORE_HEADER)] != CORE_HEADER:
            fail(f"{sheet}: header core changed: {hdr[:10]}")
        if CAPACITY_COL not in hdr:
            fail(f"{sheet}: missing {CAPACITY_COL!r}")
        ix = {h: i for i, h in enumerate(hdr)}

        n_rows = n_india = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in r):
                continue
            n_rows += 1
            uid = str(r[ix["GEM unit ID"]])
            seen_ids[uid] += 1
            status = str(r[ix["Unit status"]]).strip()
            if status not in STATUSES:
                anomalies[f"{sheet}: unexpected status {status!r}"] += 1
                continue
            country = str(r[ix["Country/area"]]).strip()
            if country != "India":
                continue
            n_india += 1
            cap = r[ix[CAPACITY_COL]]
            if not isinstance(cap, (int, float)):
                anomalies[f"{sheet}: India unit with non-numeric capacity"] += 1
                continue
            india_units.append({"route": route, "status": status, "ttpa": float(cap)})
        report.append(f"{sheet}: {n_rows} rows, {n_india} India")

    dupes = {k: v for k, v in seen_ids.items() if v > 1}
    if dupes:
        fail(f"unit ids not unique: {list(dupes)[:5]}")
    report.append(f"unit ids: {sum(seen_ids.values())} all unique")
    for k, v in anomalies.items():
        report.append(f"excluded: {k} x{v}")

    # aggregate -------------------------------------------------------------
    # measure -> route -> ttpa; routes with zero units simply do not appear.
    agg: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    counts: dict[str, int] = Counter()
    for u in india_units:
        for measure, statuses in MEASURES.items():
            if u["status"] in statuses:
                agg[measure][u["route"]] += u["ttpa"]
                agg[measure]["total"] += u["ttpa"]
                counts[measure] += 1

    report.append("")
    for m in MEASURES:
        parts = ", ".join(f"{r}={v:,.0f}" for r, v in sorted(agg[m].items()))
        report.append(f"India {m}: {counts[m]} units; ttpa: {parts}")

    # emit ------------------------------------------------------------------
    caveat = (
        "Snapshot at the named release, not a historical series: the tracker "
        "records each unit's current capacity, and start years are unknown for "
        "most Indian units, so no year-by-year reconstruction is published. "
        "The tracker covers only plants of 0.5 million tonnes per annum and "
        "greater, so small plants are outside these totals. "
        f"Source: {CITATION}, CC BY 4.0."
    )
    method = {
        "steel-capacity-operating": (
            "Sum of current capacity over Indian steelmaking furnace units with "
            "status operating or operating pre-retirement, across electric arc, "
            "basic oxygen, induction and open hearth routes. " + caveat
        ),
        "steel-capacity-construction": (
            "Sum of current capacity over Indian steelmaking furnace units with "
            "status construction. " + caveat
        ),
        "steel-capacity-announced": (
            "Sum of current capacity over Indian steelmaking furnace units with "
            "status announced. Announced projects may never be built. " + caveat
        ),
    }
    route_names = {"eaf": "electric arc", "bof": "basic oxygen", "if": "induction", "ohf": "open hearth"}
    ind_rows = [
        ("steel-capacity-operating", "Steel capacity, operating", 10),
        ("steel-capacity-construction", "Steel capacity, under construction", 40),
        ("steel-capacity-announced", "Steel capacity, announced", 50),
    ]
    for route, label in route_names.items():
        if agg["operating"].get(route):
            slug = f"steel-capacity-operating-{route}"
            ind_rows.append((slug, f"Steel capacity, operating, {label} route", 20))
            method[slug] = (
                f"Sum of current capacity over Indian {label} furnace units with "
                "status operating or operating pre-retirement. " + caveat
            )

    def q(s: str) -> str:
        return '"' + s.replace('"', '""') + '"'

    with (out / "gist-indicators.csv").open("w") as f:
        f.write("id,name,unit,category,methodology\n")
        for slug, name, _ in ind_rows:
            f.write(f"{slug},{q(name)},ttpa,Industry,{q(method[slug])}\n")

    with (out / "gist-indicator_values.csv").open("w") as f:
        f.write(
            "indicator,state,year,value,source_title,source_url,reporting_period,reporting_org,notes,verified_on\n"
        )
        def emit(slug: str, value: float, note: str = "") -> None:
            f.write(
                f"{slug},India,{RELEASE_YEAR},{value:.0f},{q(CITATION)},{SOURCE_URL},"
                f"{q(RELEASE + ' release')},Global Energy Monitor,{q(note)},{VERIFIED_ON}\n"
            )

        emit("steel-capacity-operating", agg["operating"]["total"],
             "Includes units operating while slated for retirement.")
        for route in route_names:
            v = agg["operating"].get(route)
            if v:
                emit(f"steel-capacity-operating-{route}", v)
        emit("steel-capacity-construction", agg["construction"]["total"])
        emit("steel-capacity-announced", agg["announced"]["total"],
             "Announced projects may never be built.")

    (out / "gist-report.txt").write_text("\n".join(report) + "\n")
    print("\n".join(report))
    print(f"\nwrote {out}/gist-indicators.csv, gist-indicator_values.csv, gist-report.txt")


if __name__ == "__main__":
    main()
