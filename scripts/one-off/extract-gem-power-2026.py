#!/usr/bin/env python3
"""Transform four GEM asset registers into Abhilekh inbox rows.

Usage:
  python3 extract-gem-power-2026.py --solar A.xlsx --wind B.xlsx \
      --iron-units C.xlsx --iron-plants D.xlsx [--out DIR]

Solar and wind become the first MULTI-YEAR indicator series in the archive.
Everything published before this was a snapshot. Their shape is a cumulative
stock: for each geography and each year, the summed capacity of assets whose
recorded commissioning year is that year or earlier.

That method has two consequences, stated in every methodology it produces:

1. An asset with no recorded commissioning year is absent from EVERY year, so
   the series understates the operating fleet by exactly the undated share.
   The share is computed here and written into the text, so it cannot drift
   away from the data.
2. The series reconstructs history from the fleet operating at the release
   date. An asset commissioned and later retired is invisible. This is stated
   rather than corrected, because the source carries no retired Indian rows to
   correct it with.

A geography's series starts at its own first recorded commissioning year, not
at the dataset minimum. Emitting zeros back to the dataset minimum would
assert "this state had no capacity", which the source does not support: GEM
tracks assets above a size threshold, so silence before the first tracked
asset is absence of record, not absence of capacity.

Datasets (all Global Energy Monitor; raw files stay OUTSIDE the repository,
recorded in data/raw/gem/MANIFEST.md):

- Global Solar Power Tracker, February 2026. Utility-scale (1 MW+) tab.
  Strongest file in the batch: every Indian row carries a state, and 93 per
  cent of operating capacity carries a commissioning year.
- Global Wind Power Tracker, February 2026. Weaker: 70 per cent of operating
  capacity is dated. Published with the coverage stated, because the undated
  share is spread across states (18 to 52 per cent) rather than concentrated
  in the oldest ones, so the curve's shape survives even though its level
  undercounts.
- Global Iron and Steel Tracker, June 2026 (V1), iron unit workbook. Blast
  furnaces and direct reduced iron furnaces, published as SNAPSHOTS, not as a
  series. Neither sheet carries a state, so state is joined from the
  plant-level workbook on GEM plant ID (which resolves for every Indian iron
  unit). The series was built first and withdrawn; see do_iron for why.

Ironmaking capacity is a DIFFERENT measure from the steelmaking capacity
already published from the same tracker. Blast furnaces make iron; the earlier
figures are furnace-level and plant-level crude steel. Neither is corrected
toward the other and each methodology names its own basis.

Note for the next release: the iron workbook writes the literal string
"unknown" in date cells that the solar and wind trackers leave empty. Any new
GEM file must be checked for its own sentinel before its coverage is trusted.

Verdicts recorded for the other two files in this batch, with no output:

- Global Methane Emissions Tracker V3 (December 2025). India rows exist in all
  six data sheets, but NONE carries an Indian subnational unit and no sheet
  carries a commissioning or opening year, so nothing can be placed on either
  axis the archive needs. Coal mines hold only latitude and longitude, and
  resolving 543 points to states would require a boundary file this project
  does not have and would introduce error the source does not contain. The
  file's populated numeric column for India is GEM's own modelled methane
  estimate rather than a measured quantity, and "Coal Output (Annual, Mst)" is
  empty for every Indian row.
- LNG Carrier Tracker (December 2025). 1,143 vessels, zero Indian shipowners
  and zero Indian shipbuilding yards. Vessels are mobile assets and carry no
  state geography in any case.
"""

import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

OUT_IND = "id,name,unit,category,methodology\n"
OUT_VAL = (
    "indicator,state,year,value,source_title,source_url,reporting_period,"
    "reporting_org,notes,verified_on\n"
)
VERIFIED_ON = "2026-08-08"
ORG = "Global Energy Monitor"

# Every name the states table holds. A state name that is not in this list is
# a FATAL error, never a silently dropped row: load-inbox.ts would skip it
# with a warning nobody reads, and the national total would then disagree with
# the sum of its states for a reason no page explains.
KNOWN_STATES = {
    "Andaman and Nicobar Islands", "Andhra Pradesh", "Arunachal Pradesh", "Assam",
    "Bihar", "Chandigarh", "Chhattisgarh", "Dadra and Nagar Haveli",
    "Dadra and Nagar Haveli and Daman and Diu", "Daman and Diu", "Delhi", "Goa",
    "Gujarat", "Haryana", "Himachal Pradesh", "Jammu and Kashmir", "Jharkhand",
    "Karnataka", "Kerala", "Ladakh", "Lakshadweep", "Madhya Pradesh", "Maharashtra",
    "Manipur", "Meghalaya", "Mizoram", "Nagaland", "Odisha", "Puducherry", "Punjab",
    "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana", "Tripura", "Uttar Pradesh",
    "Uttarakhand", "West Bengal",
}


# The iron workbook writes the literal string "unknown" where the solar and
# wind trackers leave a cell empty. Treating that as a value counted every row
# as carrying a retired date, so the sentinel set is shared by every check.
SENTINELS = {"", "unknown", "n/a", "na", "none", "tbd", "not found", "-"}


def blank(v) -> bool:
    return v is None or str(v).strip().lower() in SENTINELS


def q(s: str) -> str:
    return '"' + str(s).replace('"', '""') + '"'


def arg(name: str) -> str | None:
    return sys.argv[sys.argv.index(name) + 1] if name in sys.argv else None


def open_sheet(path: str, sheet: str, required: list[str]):
    """Open a sheet and prove its header carries every column we read.

    Proving the header verbatim is the whole point: a GEM release that renames
    or reorders a column must stop this script, not quietly shift a column
    index and publish the wrong number under a real source citation.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"FATAL: {Path(path).name}: no sheet {sheet!r} (has {wb.sheetnames})")
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    while hdr is not None and all(c is None for c in hdr):
        hdr = next(rows)
    names = [str(c) if c is not None else "" for c in hdr]
    missing = [c for c in required if c not in names]
    if missing:
        sys.exit(f"FATAL: {Path(path).name}/{sheet}: missing column(s) {missing}")
    return wb, rows, {h: i for i, h in enumerate(names)}


def year_of(raw) -> int | None:
    """Commissioning year from a float (2005.0), an int, or a string ('2003')."""
    if blank(raw):
        return None
    text = str(raw).strip()[:4]
    try:
        y = int(text)
    except ValueError:
        return None
    return y if 1900 <= y <= 2100 else None


def number(raw) -> float | None:
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def pct(part: float, whole: float) -> str:
    return f"{100 * part / whole:.1f}" if whole else "0.0"


def fmt(v: float) -> str:
    """Trim a float that is really an integer, so CSV values stay readable."""
    return f"{v:.0f}" if abs(v - round(v)) < 1e-9 else f"{v:.1f}"


def grouped(v: float) -> str:
    """Same number for prose, with thousands separators."""
    return f"{v:,.0f}" if abs(v - round(v)) < 1e-9 else f"{v:,.1f}"


class Series:
    """Cumulative-stock builder for one indicator.

    Assets are added with a geography, a commissioning year and a capacity.
    Undated assets are counted but never contribute to a year, which is the
    undercount the methodology has to declare.
    """

    def __init__(self, indicator_id: str, name: str, unit: str, category: str):
        self.id = indicator_id
        self.name = name
        self.unit = unit
        self.category = category
        self.dated: dict[str, list[tuple[int, float]]] = defaultdict(list)
        self.total_units = 0
        self.total_cap = 0.0
        self.dated_units = 0
        self.dated_cap = 0.0
        self.no_state_dated_cap = 0.0
        self.no_state_units = 0

    def add(self, state: str | None, year: int | None, cap: float) -> None:
        self.total_units += 1
        self.total_cap += cap
        if year is None:
            return
        self.dated_units += 1
        self.dated_cap += cap
        # The national series is built from every dated asset, including those
        # whose state is blank, so it is more complete than the sum of states.
        self.dated["India"].append((year, cap))
        if state is None:
            self.no_state_units += 1
            self.no_state_dated_cap += cap
        else:
            self.dated[state].append((year, cap))

    def points(self):
        """(geography, year, cumulative value, contributing asset count)."""
        out = []
        for geo, assets in self.dated.items():
            assets.sort()
            last = max(y for y, _ in assets)
            running, count, i = 0.0, 0, 0
            for y in range(min(y for y, _ in assets), last + 1):
                while i < len(assets) and assets[i][0] <= y:
                    running += assets[i][1]
                    count += 1
                    i += 1
                out.append((geo, y, running, count))
        return out


def emit(series: Series, methodology: str, source_title: str, source_url: str,
         ind_lines: list[str], val_lines: list[str]) -> None:
    ind_lines.append(
        ",".join([series.id, q(series.name), q(series.unit), q(series.category),
                  q(methodology)]) + "\n"
    )
    pts = series.points()
    for geo, year, value, count in sorted(pts, key=lambda p: (p[0] != "India", p[0], p[1])):
        note = f"Cumulative over {count} unit{'s' if count != 1 else ''} commissioned by end {year}."
        val_lines.append(
            ",".join([
                series.id, q(geo), str(year), fmt(value), q(source_title), source_url,
                q(f"End of {year}"), q(ORG), q(note), VERIFIED_ON,
            ]) + "\n"
        )
    states = {g for g, *_ in pts} - {"India"}
    bad = states - KNOWN_STATES
    if bad:
        sys.exit(f"FATAL: {series.id}: state name(s) not in the states table: {sorted(bad)}")
    print(
        f"  {series.id}: {len(pts)} values across {len(states)} states plus India, "
        f"{series.dated_units}/{series.total_units} units dated "
        f"({pct(series.dated_cap, series.total_cap)}% of capacity)"
    )


# --- solar -------------------------------------------------------------------

SOLAR_SHEET = "Utility-Scale (1 MW+)"
SOLAR_COLS = ["Country/Area", "Capacity (MW)", "Capacity Rating", "Status",
              "Start year", "Retired year", "State/Province", "Other IDs (unit/phase)"]


def do_solar(path: str, ind: list[str], val: list[str]) -> None:
    wb, rows, ix = open_sheet(path, SOLAR_SHEET, SOLAR_COLS)
    s = Series("solar-capacity-operating", "Utility-scale solar capacity, operating",
               "MW", "Energy")
    rating = defaultdict(float)
    retired = tz = wksl = 0
    for r in rows:
        if r[ix["Country/Area"]] != "India":
            continue
        if not blank(r[ix["Retired year"]]):
            retired += 1
        if str(r[ix["Status"]]).strip().lower() != "operating":
            continue
        cap = number(r[ix["Capacity (MW)"]])
        if cap is None:
            sys.exit(f"FATAL: solar: operating Indian row with unreadable capacity: {r[:4]}")
        rating[str(r[ix["Capacity Rating"]]).strip()] += cap
        other = str(r[ix["Other IDs (unit/phase)"]] or "")
        if "TZ" in other:
            tz += 1
        elif "WKSL" in other:
            wksl += 1
        st = r[ix["State/Province"]]
        s.add(None if blank(st) else str(st).strip(), year_of(r[ix["Start year"]]), cap)
    wb.close()
    total = sum(rating.values())
    method = (
        "Cumulative nameplate capacity of utility-scale solar projects, being those of "
        "1 MW and above, with status operating, summed over the projects whose recorded "
        "commissioning year is the stated year or earlier. "
        f"Of {s.total_units} operating Indian projects {s.dated_units} carry a commissioning "
        f"year, being {grouped(s.dated_cap)} MW of {grouped(s.total_cap)} MW "
        f"({pct(s.dated_cap, s.total_cap)} per cent of operating capacity); the undated "
        "remainder is absent from every year, so each value understates the fleet by that "
        "much. Global Energy Monitor records capacity as either an alternating-current "
        "rating or a direct-current peak rating and frequently does not state which: of "
        f"operating Indian capacity {pct(rating.get('MWac', 0), total)} per cent is stated "
        f"MWac, {pct(rating.get('MWp/dc', 0), total)} per cent MWp/dc and "
        f"{pct(rating.get('unknown', 0), total)} per cent unspecified. The ratings are summed "
        "as published and are not converted between bases. The series is reconstructed from "
        "the fleet operating at the release date, so a project commissioned and later "
        "retired does not appear; the source records no retired Indian projects. Records for "
        f"{tz} Indian projects derive partly or wholly from the TransitionZero Solar Asset "
        f"Mapper, distributed under CC BY-NC 4.0, and {wksl} from the wiki-solar.org dataset."
    )
    if retired:
        sys.exit(f"FATAL: solar: {retired} Indian rows carry a retired year; the "
                 "no-retirements claim in the methodology is no longer true")
    emit(s, method, "Global Energy Monitor, Global Solar Power Tracker, February 2026 release",
         "https://globalenergymonitor.org/projects/global-solar-power-tracker/", ind, val)


# --- wind --------------------------------------------------------------------

WIND_COLS = ["Country/Area", "Capacity (MW)", "Status", "Start year", "Retired year",
             "State/Province", "Installation Type"]


def do_wind(path: str, ind: list[str], val: list[str]) -> None:
    wb, rows, ix = open_sheet(path, "Data", WIND_COLS)
    s = Series("wind-capacity-operating", "Utility-scale wind capacity, operating",
               "MW", "Energy")
    retired = 0
    for r in rows:
        if r[ix["Country/Area"]] != "India":
            continue
        if not blank(r[ix["Retired year"]]):
            retired += 1
        if str(r[ix["Status"]]).strip().lower() != "operating":
            continue
        cap = number(r[ix["Capacity (MW)"]])
        if cap is None:
            sys.exit(f"FATAL: wind: operating Indian row with unreadable capacity: {r[:4]}")
        st = r[ix["State/Province"]]
        s.add(None if blank(st) else str(st).strip(), year_of(r[ix["Start year"]]), cap)
    wb.close()
    method = (
        "Cumulative nameplate capacity of wind projects with status operating, summed over "
        "the projects whose recorded commissioning year is the stated year or earlier. The "
        "tracker's global threshold excludes wind farms below 10 MW, so small installations "
        "are outside the series entirely. "
        f"COVERAGE IS PARTIAL: of {s.total_units} operating Indian projects only "
        f"{s.dated_units} carry a commissioning year, being {grouped(s.dated_cap)} MW of "
        f"{grouped(s.total_cap)} MW ({pct(s.dated_cap, s.total_cap)} per cent of operating "
        "capacity). Every value therefore understates installed wind capacity, and the "
        "latest value should not be read as India's wind total. The undated share is spread "
        "across states rather than concentrated in the earliest ones, so the shape of the "
        "series is less affected than its level. "
        # Only claim a state-to-national discrepancy when one exists. India's
        # single stateless wind project is also undated, so it falls out of
        # both the state series and the national one and the two do reconcile.
        + (
            f"{fmt(s.no_state_dated_cap)} MW of dated capacity sits in {s.no_state_units} "
            "project(s) with no recorded state, which the national series includes and no "
            "state series does, so the state values do not sum to the national one. "
            if s.no_state_units else
            "Every dated project carries a state, so the state values sum to the national "
            "one. "
        )
        + "The series is reconstructed from the fleet operating at the release date, so a "
        "project commissioned and later retired does not appear; the source records no "
        "retired Indian projects."
    )
    if retired:
        sys.exit(f"FATAL: wind: {retired} Indian rows carry a retired year; the "
                 "no-retirements claim in the methodology is no longer true")
    emit(s, method, "Global Energy Monitor, Global Wind Power Tracker, February 2026 release",
         "https://globalenergymonitor.org/projects/global-wind-power-tracker/", ind, val)


# --- iron --------------------------------------------------------------------

# Both statuses are producing today. "mothballed" is disused but not
# dismantled, so it is not operating capacity and stays out.
IRON_OPERATING = {"operating", "operating pre-retirement"}

# The last field is why no year-by-year series is published, and it differs by
# sheet. Relining is a blast-furnace concept and the blast-furnace sheet is the
# only one that records it, so only that sheet may cite it.
IRON_SHEETS = [
    ("Blast furnaces", "iron-capacity-bf-operating",
     "Ironmaking capacity, blast furnace route, operating",
     "blast furnaces, which reduce iron ore with coke to make pig iron",
     "every dated operating Indian unit carries a recorded relining, so attributing today's "
     "rating to the year a furnace first lit would state a figure the source does not support"),
    ("Direct reduced iron furnaces", "iron-capacity-dri-operating",
     "Ironmaking capacity, direct reduced iron route, operating",
     "direct reduced iron furnaces, which reduce iron ore without melting it",
     "a unit's rating can change over its life while the workbook keeps only the current "
     "figure, with the start year missing for a fifth of operating capacity"),
]
IRON_COLS = ["GEM plant ID", "Country/area", "Unit status", "Start date",
             "Current capacity (ttpa)", "Retired date"]


def plant_states(path: str) -> dict[str, str]:
    wb, rows, ix = open_sheet(path, "Plant data", ["GEM plant ID", "Subnational unit"])
    country = next(k for k in ix if "countr" in k.lower())
    out = {}
    for r in rows:
        if r[ix[country]] != "India":
            continue
        st = r[ix["Subnational unit"]]
        if st not in (None, ""):
            out[str(r[ix["GEM plant ID"]]).strip()] = str(st).strip()
    wb.close()
    return out


IRON_YEAR = 2026  # the June 2026 (V1) release; a snapshot is dated by its release


def do_iron(units_path: str, plants_path: str, ind: list[str], val: list[str]) -> None:
    """Ironmaking capacity as SNAPSHOTS, deliberately not as a time series.

    The unit sheets carry a start date, so a cumulative series is mechanically
    available and was built first. It was withdrawn: "Current capacity (ttpa)"
    is a present-day rating, every dated operating Indian blast furnace has a
    recorded relining, and the oldest still runs from 1919. Summing current
    ratings by start year would put a 2018 rating on a 1919 axis and publish
    "India had 975 ttpa of blast furnace capacity in 1919" under a real
    citation. Solar and wind escape this because Global Energy Monitor tracks
    project PHASES: an expansion is a new row with its own year and its own
    capacity, so each figure belongs to the year it is filed under.
    """
    p2s = plant_states(plants_path)
    print(f"  plant-level join table: {len(p2s)} Indian plants with a state")
    for sheet, ind_id, name, what, why_no_series in IRON_SHEETS:
        wb, rows, ix = open_sheet(units_path, sheet, IRON_COLS)
        by_state: dict[str, float] = defaultdict(float)
        by_state_n: dict[str, int] = defaultdict(int)
        units = retired = unjoined = 0
        total = 0.0
        oldest = None
        for r in rows:
            if r[ix["Country/area"]] != "India":
                continue
            if not blank(r[ix["Retired date"]]):
                retired += 1
            if str(r[ix["Unit status"]]).strip().lower() not in IRON_OPERATING:
                continue
            cap = number(r[ix["Current capacity (ttpa)"]])
            if cap is None:
                sys.exit(f"FATAL: {sheet}: operating Indian unit with unreadable capacity")
            st = p2s.get(str(r[ix["GEM plant ID"]]).strip())
            if st is None:
                unjoined += 1
                continue
            units += 1
            total += cap
            by_state[st] += cap
            by_state_n[st] += 1
            y = year_of(r[ix["Start date"]])
            if y is not None:
                oldest = y if oldest is None else min(oldest, y)
        wb.close()
        if unjoined:
            sys.exit(f"FATAL: {sheet}: {unjoined} operating Indian units did not join to a "
                     "plant state; state coverage is no longer complete")
        bad = set(by_state) - KNOWN_STATES
        if bad:
            sys.exit(f"FATAL: {ind_id}: state name(s) not in the states table: {sorted(bad)}")
        method = (
            f"Sum of current capacity over Indian {what}, with status operating or operating "
            f"pre-retirement, as at the June 2026 (V1) release. Covers {units} units, every "
            "one of which is located by joining the unit workbook to the plant-level "
            "workbook on GEM plant ID; the unit sheets carry no state of their own. This "
            "measures IRONMAKING and is a different quantity from the steelmaking capacity "
            "published from the same tracker, which is neither corrected toward this nor "
            "derived from it. No year-by-year series is published even though the workbook "
            "records a start year for each unit: the capacity figure is a present-day rating, "
            f"the oldest operating unit dates from {oldest}, and {why_no_series}."
        )
        ind.append(",".join([ind_id, q(name), q("ttpa"), q("Industry"), q(method)]) + "\n")
        title = "Global Energy Monitor, Global Iron and Steel Tracker, June 2026 (V1) release"
        url = "https://globalenergymonitor.org/projects/global-iron-and-steel-tracker/"
        for geo, value in [("India", total)] + sorted(by_state.items()):
            n = units if geo == "India" else by_state_n[geo]
            val.append(",".join([
                ind_id, q(geo), str(IRON_YEAR), fmt(value), q(title), url,
                q("June 2026 (V1) release"), q(ORG),
                q(f"{n} operating unit{'s' if n != 1 else ''}."), VERIFIED_ON,
            ]) + "\n")
        print(f"  {ind_id}: {len(by_state) + 1} snapshot values ({len(by_state)} states plus "
              f"India), {units} units, {fmt(total)} ttpa, oldest unit {oldest}")
        if retired:
            print(f"    note: {retired} Indian unit(s) in this sheet carry a retired date "
                  "and are excluded")


def main() -> None:
    out = Path(arg("--out") or ".")
    ind: list[str] = []
    val: list[str] = []
    if p := arg("--solar"):
        print("solar:")
        do_solar(p, ind, val)
    if p := arg("--wind"):
        print("wind:")
        do_wind(p, ind, val)
    if (u := arg("--iron-units")) and (pl := arg("--iron-plants")):
        print("iron:")
        do_iron(u, pl, ind, val)
    if not ind:
        sys.exit("nothing to do: pass --solar, --wind, or --iron-units with --iron-plants")
    (out / "indicators.add.csv").write_text(OUT_IND + "".join(ind))
    (out / "indicator_values.add.csv").write_text(OUT_VAL + "".join(val))
    print(f"\nwrote {len(ind)} indicators and {len(val)} values to {out}/")


if __name__ == "__main__":
    main()
